#!/usr/bin/env python3
"""
Patch seed_data.py (add Roles + MLModels sections)
and generate test_files_sandbox/ with realistic files.
Run this script on the server.
"""
import os, sys, struct, zlib

BASE = '/home/adoum/Integrated_BI/sotifibre_backend_django'
SANDBOX = os.path.join(BASE, 'test_files_sandbox')
os.makedirs(SANDBOX, exist_ok=True)

# ============================================================
# STEP 1: Patch seed_data.py
# ============================================================
SEED_PATH = os.path.join(BASE, 'seed_data.py')
with open(SEED_PATH) as f:
    seed = f.read()

# --- 1a: Add missing imports after existing imports block ---
OLD_IMPORTS = "from apps.notifications.models import Notification"
NEW_IMPORTS = (
    "from apps.notifications.models import Notification\n"
    "from apps.users.models import Role\n"
    "from apps.ml_analytics.models import MLModel, ModelTrainingLog\n"
    "from apps.star_schema.models import (\n"
    "    DimensionalSchema, DimensionHierarchy, CustomCalculation, FactRelationship\n"
    ")\n"
    "from apps.data_sources.models import DataSource, DataTable, Connection"
)
if "from apps.users.models import Role" not in seed:
    seed = seed.replace(OLD_IMPORTS, NEW_IMPORTS, 1)
    print("  [PATCH] Imports mis a jour")

# --- 1b: Add cleanup for new models ---
OLD_CLEANUP = "Team.objects.all().delete()"
NEW_CLEANUP = (
    "Team.objects.all().delete()\n"
    "Role.objects.all().delete()\n"
    "try:\n"
    "    MLModel.objects.all().delete()\n"
    "    ModelTrainingLog.objects.all().delete()\n"
    "except: pass\n"
    "try:\n"
    "    FactRelationship.objects.all().delete()\n"
    "    DimensionHierarchy.objects.all().delete()\n"
    "    CustomCalculation.objects.all().delete()\n"
    "    Connection.objects.all().delete()\n"
    "except: pass"
)
if "Role.objects.all().delete()" not in seed:
    seed = seed.replace(OLD_CLEANUP, NEW_CLEANUP, 1)
    print("  [PATCH] Cleanup mis a jour")

# --- 1c: Insert Roles section after Teams section ---
ROLES_SECTION = '''
# ---------------------------------------------------------
# 2b. ROLES
# ---------------------------------------------------------
print("\\n[2b/14] Roles...")

roles_spec = [
    ('superadmin',   'Administrateur Systeme',
     'Acces complet a toutes les fonctionnalites et parametres systeme.',
     ['admin:all', 'users:all', 'data:all', 'bi:all', 'ml:all']),
    ('bi_developer', 'Developpeur BI',
     'Developpement et maintenance des pipelines ETL, schemas dimensionnels et modeles ML.',
     ['pipelines:read', 'pipelines:write', 'datasources:all', 'warehouse:all', 'ml:all']),
    ('bi_analyst',   'Analyste BI',
     'Creation et gestion des tableaux de bord, KPIs, rapports et analyses de donnees.',
     ['dashboards:all', 'kpis:all', 'reports:all', 'data:read', 'ml:read']),
    ('bi_consumer',  'Utilisateur BI',
     'Consultation des tableaux de bord et rapports autorises. Pas de modification.',
     ['dashboards:read', 'kpis:read', 'reports:read']),
    ('viewer',       'Lecteur',
     'Acces en lecture seule aux ressources partageees publiquement.',
     ['dashboards:read']),
]

for rname, rdisplay, rdesc, rperms in roles_spec:
    r, c = Role.objects.get_or_create(name=rname, defaults=dict(
        display_name=rdisplay, description=rdesc, permissions=rperms,
    ))
    if c: print(f"  [OK] {rname} - {rdisplay}")

print(f"  Total roles: {Role.objects.count()}")

'''

if "2b. ROLES" not in seed:
    # Insert after the Teams section
    INSERT_AFTER = "# ---------------------------------------------------------\n# 3. SOURCES DE DONNEES"
    seed = seed.replace(INSERT_AFTER, ROLES_SECTION + INSERT_AFTER, 1)
    print("  [PATCH] Section Roles ajoutee")

# --- 1d: Insert ML Models section after Schemas section ---
ML_SECTION = '''
# ---------------------------------------------------------
# 10b. MODELES ML
# ---------------------------------------------------------
print("\\n[10b/14] Modeles ML...")

try:
    ml_specs = [
        dict(
            name='ML_Prevision_Bande_Passante',
            description='Modele de prevision de la consommation bande passante par noeud fibre pour les 30 prochains jours. Entraine sur 12 mois de donnees horaires IoT.',
            model_type='forecast', algorithm='prophet', status='deployed', version=3,
            dimensional_schema=schema_cons,
            measure=measures['Data Transferee Totale'],
            features=['date', 'noeud_id', 'debit_descendant_mbps', 'taux_utilisation_pct', 'jour_semaine', 'heure'],
            parameters={'seasonality_mode': 'multiplicative', 'changepoint_prior_scale': 0.05, 'horizon_days': 30},
            accuracy=94.2, rmse=12.4, mae=8.7, mape=3.8,
            training_data_size=365000, training_duration_ms=84200,
            last_trained=None, training_frequency='monthly',
            is_active=True, tags=['prevision', 'bande-passante', 'prophet'],
            owner=dev_user, team=team_data,
        ),
        dict(
            name='ML_Detection_Anomalies_Reseau',
            description='Detection en temps reel des anomalies de debit et latence sur les noeuds OLT du reseau fibre SOTIFibre. Seuil de detection : 3 sigmas.',
            model_type='anomaly', algorithm='isolation_forest', status='deployed', version=2,
            dimensional_schema=schema_cons,
            measure=measures['Debit Descendant Moyen'],
            features=['debit_descendant_mbps', 'debit_montant_mbps', 'latence_ms', 'taux_utilisation_pct'],
            parameters={'contamination': 0.05, 'n_estimators': 200, 'max_samples': 'auto'},
            accuracy=91.8, precision=89.4, recall=94.2, f1_score=91.7,
            training_data_size=1240000, training_duration_ms=42800,
            last_trained=None, training_frequency='weekly',
            is_active=True, tags=['anomalie', 'reseau', 'isolation-forest'],
            owner=dev_user, team=team_data,
        ),
        dict(
            name='ML_Segmentation_Clients_Fibre',
            description='Segmentation des abonnes SOTIFibre en 5 clusters selon le profil de consommation : intensite, regularite, pic horaire et zone geographique.',
            model_type='segmentation', algorithm='kmeans', status='trained', version=1,
            dimensional_schema=schema_cons,
            measure=measures['Data Transferee Totale'],
            features=['data_transferee_gb', 'debit_moyen_mbps', 'heure_pic', 'wilaya', 'type_offre'],
            parameters={'n_clusters': 5, 'init': 'k-means++', 'max_iter': 300, 'n_init': 10},
            accuracy=88.6,
            training_data_size=42380, training_duration_ms=18900,
            last_trained=None, training_frequency='monthly',
            is_active=True, tags=['segmentation', 'clients', 'kmeans'],
            owner=analyst_user, team=team_bi,
        ),
        dict(
            name='ML_Classification_Incidents_Reseau',
            description='Classification automatique des incidents reseau par type (coupure, degradation, securite) et severite pour priorisation des interventions.',
            model_type='classification', algorithm='random_forest', status='deployed', version=4,
            dimensional_schema=schema_inc,
            measure=measures['Nombre Incidents'],
            features=['type_equipement', 'zone_id', 'heure', 'jour_semaine', 'duree_precedente', 'nb_clients_zone'],
            parameters={'n_estimators': 150, 'max_depth': 12, 'min_samples_split': 5, 'class_weight': 'balanced'},
            accuracy=93.7, precision=92.1, recall=94.8, f1_score=93.4, roc_auc=0.971,
            training_data_size=28640, training_duration_ms=31500,
            last_trained=None, training_frequency='monthly',
            is_active=True, tags=['classification', 'incidents', 'random-forest'],
            owner=dev_user, team=team_data,
        ),
        dict(
            name='ML_Regression_Cout_Interventions',
            description='Modele de regression estimant le cout et la duree des interventions terrain avant dispatch technicien, optimisant la planification des equipes.',
            model_type='regression', algorithm='gradient_boosting', status='trained', version=2,
            dimensional_schema=schema_int,
            measure=measures['Cout Total Interventions'],
            features=['type_intervention', 'zone_id', 'type_equipement', 'technicien_experience', 'heure'],
            parameters={'n_estimators': 200, 'learning_rate': 0.05, 'max_depth': 5, 'subsample': 0.8},
            accuracy=87.3, rmse=1240.5, mae=890.2, mape=8.4,
            training_data_size=18720, training_duration_ms=22100,
            last_trained=None, training_frequency='quarterly',
            is_active=True, tags=['regression', 'interventions', 'gradient-boosting'],
            owner=analyst_user, team=team_bi,
        ),
    ]

    from django.utils import timezone
    from datetime import timedelta

    ml_models = {}
    mcount = 0
    for mspec in ml_specs:
        name = mspec['name']
        mspec['last_trained'] = timezone.now() - timedelta(days=7)
        m, c = MLModel.objects.get_or_create(name=name, defaults=mspec)
        ml_models[name] = m
        if c:
            print(f"  [OK] {name}")
            mcount += 1

    print(f"  Total modeles ML: {mcount} crees")

    # Training logs for each model
    tl_count = 0
    for mname, m in ml_models.items():
        if not ModelTrainingLog.objects.filter(model=m).exists():
            for i in range(3):
                started = timezone.now() - timedelta(days=7 + i * 30)
                dur = m.training_duration_ms or 30000
                ModelTrainingLog.objects.create(
                    model=m,
                    status='success' if i > 0 else 'success',
                    started_at=started,
                    completed_at=started + timedelta(milliseconds=dur),
                    duration_ms=dur,
                    accuracy=m.accuracy,
                    training_data_size=m.training_data_size or 10000,
                    notes=f'Entrainement v{m.version - i} - donnees du mois precedent.',
                )
                tl_count += 1
    print(f"  [OK] {tl_count} logs d entrainement crees")

except Exception as e:
    import traceback
    print(f"  [WARN] MLModels: {e}")
    traceback.print_exc()

# ---------------------------------------------------------
# 10c. SCHEMAS ETOILE - HIERARCHIES ET CALCULS
# ---------------------------------------------------------
print("\\n[10c/14] Hierarchies et calculs dimensionnels...")

try:
    hier_specs = [
        (schema_inc, 'Hierarchie Temporelle Incidents',
         'Hierarchie temps : Annee > Trimestre > Mois > Semaine > Jour',
         ['annee', 'trimestre', 'mois', 'semaine', 'jour'], dim_temps_tbl),
        (schema_inc, 'Hierarchie Geographique Algerie',
         'Hierarchie geographique : Region > Wilaya > Daira > Commune',
         ['region', 'wilaya', 'daira', 'commune'], dim_loc_tbl),
        (schema_cons, 'Hierarchie Temporelle Performance',
         'Hierarchie pour analyse performance : Annee > Mois > Semaine > Jour > Heure',
         ['annee', 'mois', 'semaine', 'jour', 'heure'], dim_temps_tbl),
        (schema_int, 'Hierarchie Equipements Reseau',
         'Hierarchie reseau : Region > Zone > Noeud > Equipement',
         ['region', 'zone_reseau', 'noeud', 'equipement'], dim_equip_tbl),
    ]

    hcount = 0
    for schema, hname, hdesc, hlevels, dim_tbl in hier_specs:
        h, c = DimensionHierarchy.objects.get_or_create(
            schema=schema, name=hname,
            defaults=dict(description=hdesc, levels=hlevels,
                          dimension_table=dim_tbl, is_active=True),
        )
        if c:
            print(f"  [OK] {hname}")
            hcount += 1

    calc_specs = [
        (schema_inc, 'Taux Indisponibilite Reseau',
         'Calcul du taux d indisponibilite reseau en complement du taux de resolution.',
         '(1 - taux_resolution_pct / 100) * 100', '%', 2),
        (schema_cons, 'Bande Passante Residuelle',
         'Capacite bande passante non utilisee calculee depuis le taux d utilisation.',
         '(1 - taux_utilisation_pct / 100) * capacite_totale_mbps', 'Mbps', 1),
        (schema_cons, 'Ratio Montant Descendant',
         'Rapport entre le debit montant et descendant - indicateur d equilibre du trafic.',
         'debit_montant_mbps / NULLIF(debit_descendant_mbps, 0)', 'Ratio', 3),
        (schema_int, 'Cout Moyen par Heure Intervention',
         'Cout horaire des interventions terrain pour optimisation allocation ressources.',
         'cout_total_da / NULLIF(duree_intervention_h, 0)', 'DA/h', 0),
    ]

    ccount = 0
    for schema, cname, cdesc, cformula, cunit, cdec in calc_specs:
        c_obj, c = CustomCalculation.objects.get_or_create(
            schema=schema, name=cname,
            defaults=dict(description=cdesc, formula=cformula,
                          output_unit=cunit, decimal_places=cdec, is_active=True),
        )
        if c:
            print(f"  [OK] Calcul: {cname}")
            ccount += 1

    print(f"  [OK] {hcount} hierarchies + {ccount} calculs crees")

except Exception as e:
    import traceback
    print(f"  [WARN] Hierarchies/Calculs: {e}")
    traceback.print_exc()

# ---------------------------------------------------------
# 10d. CONNEXIONS DB
# ---------------------------------------------------------
print("\\n[10d/14] Connexions DB...")
try:
    conn_specs = [
        dict(name='CONN_Production_PostgreSQL',
             description='Connexion directe a la base PostgreSQL de production SOTIFibre.',
             connection_type='postgresql',
             host='192.168.224.128', port=5432, database='sotifibre_db',
             username='sotifibre_readonly', password='readonly_2026!',
             is_active=True, last_tested=None,
             owner=dev_user, tags=['production', 'postgresql']),
        dict(name='CONN_DW_PostgreSQL',
             description='Connexion a l entrepot de donnees - schema sotifibre_dw.',
             connection_type='postgresql',
             host='192.168.224.128', port=5432, database='supervisionolt_db',
             username='supervisionolt_admin', password='dw_2026!',
             is_active=True, last_tested=None,
             owner=dev_user, tags=['datawarehouse', 'postgresql']),
    ]
    conn_count = 0
    for cspec in conn_specs:
        name = cspec['name']
        import django.utils.timezone as tz
        cspec['last_tested'] = tz.now()
        c_obj, c = Connection.objects.get_or_create(name=name, defaults=cspec)
        if c:
            print(f"  [OK] {name}")
            conn_count += 1
    print(f"  Total connexions: {conn_count}")
except Exception as e:
    print(f"  [WARN] Connexions: {e}")

'''

if "10b. MODELES ML" not in seed:
    # Insert after schema_interventions_techniques creation (10. SCHEMAS DIMENSIONNELS section end)
    INSERT_AFTER_ML = "# ---------------------------------------------------------\n# 11. TABLEAUX DE BORD"
    seed = seed.replace(INSERT_AFTER_ML, ML_SECTION + INSERT_AFTER_ML, 1)
    print("  [PATCH] Section MLModels + Hierarchies ajoutee")

# --- 1e: Update final summary ---
OLD_SUMMARY = '    print("=" * 60 + "\\n")'
NEW_SUMMARY = (
    '    print("=" * 60 + "\\n")\n\n'
    'try:\n'
    '    print(f"  Roles           : {Role.objects.count()}")\n'
    '    print(f"  Modeles ML      : {MLModel.objects.count()}")\n'
    '    print(f"  Logs training   : {ModelTrainingLog.objects.count()}")\n'
    '    print(f"  Hierarchies     : {DimensionHierarchy.objects.count()}")\n'
    '    print(f"  Calculs dim.    : {CustomCalculation.objects.count()}")\n'
    'except: pass\n'
)
if "Roles           :" not in seed:
    seed = seed.replace(OLD_SUMMARY, NEW_SUMMARY, 1)
    print("  [PATCH] Resume final mis a jour")

with open(SEED_PATH, 'w') as f:
    f.write(seed)
print("[OK] seed_data.py mis a jour")

# ============================================================
# STEP 2: Generate test_files_sandbox/
# ============================================================
print("\n[GEN] Generation des fichiers sandbox...")

# --- 2a: logs_reseau_24h.csv ---
import csv, io
rows = []
rows.append(['id', 'noeud_id', 'horodatage', 'debit_montant_mbps', 'debit_descendant_mbps',
             'taux_utilisation_pct', 'data_transferee_gb', 'latence_ms'])
import random
random.seed(42)
noeuds = ['OLT-ALGER-01', 'OLT-ALGER-02', 'OLT-ALGER-03', 'OLT-ORAN-01',
          'OLT-CONSTANTINE-01', 'OLT-ANNABA-01', 'OLT-BLIDA-01']
base_time = '2026-05-15 00:00:00'
from datetime import datetime, timedelta
t0 = datetime(2026, 5, 15, 0, 0, 0)
rec_id = 1
for h in range(24):
    for noeud in noeuds:
        t = t0 + timedelta(hours=h)
        base_debit = random.uniform(120, 480)
        montant = round(base_debit * 0.35 + random.uniform(-20, 20), 2)
        descendant = round(base_debit + random.uniform(-30, 30), 2)
        util = round(min(99.9, descendant / 600 * 100), 2)
        data_gb = round(descendant * 3600 / 1024 / 8, 3)
        latence = random.randint(2, 18)
        rows.append([rec_id, noeud, t.strftime('%Y-%m-%d %H:%M:%S'),
                     montant, descendant, util, data_gb, latence])
        rec_id += 1

csv_path = os.path.join(SANDBOX, 'logs_reseau_24h.csv')
with open(csv_path, 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f, delimiter=';')
    writer.writerows(rows)
print(f"  [OK] logs_reseau_24h.csv ({rec_id-1} lignes)")

# --- 2b: clients_fibre.json ---
import json
wilayas = ['Alger', 'Oran', 'Constantine', 'Annaba', 'Blida', 'Setif', 'Tlemcen', 'Bejaia']
offres = [1, 2, 3]  # offre_id
statuts = ['actif', 'actif', 'actif', 'suspendu', 'resilié']
prenoms = ['Mohamed', 'Ahmed', 'Fatima', 'Nadia', 'Karim', 'Sara', 'Rachid', 'Amina', 'Youcef', 'Meriem']
noms = ['Benali', 'Boudiaf', 'Hamidi', 'Meziane', 'Brahim', 'Khaled', 'Amrani', 'Larbi', 'Saadi', 'Ferhat']
clients_data = []
for i in range(50):
    random.seed(i * 7)
    d_act = (t0 - timedelta(days=random.randint(30, 1800))).strftime('%Y-%m-%d')
    clients_data.append({
        'id': i + 1,
        'nom': random.choice(noms),
        'prenom': random.choice(prenoms),
        'email': f"client{i+1:03d}@sotifibre-client.dz",
        'telephone': f"0{random.choice([5,6,7])}{random.randint(10000000, 99999999)}",
        'wilaya': random.choice(wilayas),
        'offre_id': random.choice(offres),
        'date_activation': d_act,
        'statut': random.choice(statuts),
    })
json_path = os.path.join(SANDBOX, 'clients_fibre.json')
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump({'count': len(clients_data), 'source': 'SRC_CRM_Clients_CSV', 'data': clients_data}, f, ensure_ascii=False, indent=2)
print(f"  [OK] clients_fibre.json ({len(clients_data)} clients)")

# --- 2c: rapport_incidents_mensuel.xlsx ---
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Sheet 1: Synthese
ws1 = wb.active
ws1.title = 'Synthese Mensuelle'
ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 20

header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
alt_fill = PatternFill(start_color='EEF2F7', end_color='EEF2F7', fill_type='solid')
header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
title_font = Font(name='Calibri', bold=True, size=14, color='1E3A5F')
thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

ws1.merge_cells('A1:D1')
ws1['A1'] = 'SOTIFibre - Rapport Mensuel Incidents Reseau - Mai 2026'
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 30

ws1.merge_cells('A2:D2')
ws1['A2'] = 'Periode : 01/05/2026 - 31/05/2026  |  Genere le 15/05/2026  |  Version 1.0'
ws1['A2'].alignment = Alignment(horizontal='center')
ws1['A2'].font = Font(name='Calibri', italic=True, color='666666')

ws1.append([])
headers = ['Indicateur', 'Valeur Mai 2026', 'Valeur Avril 2026', 'Evolution']
ws1.append(headers)
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=4, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin

data_rows = [
    ['Total Incidents', 312, 287, '+8.7%'],
    ['Incidents Critiques', 18, 24, '-25.0%'],
    ['Incidents Majeurs', 67, 58, '+15.5%'],
    ['Incidents Mineurs', 227, 205, '+10.7%'],
    ['Taux Resolution < 4h (%)', '91.2%', '88.6%', '+2.6 pts'],
    ['MTTR Moyen (heures)', 4.2, 5.1, '-17.6%'],
    ['Clients Impactes (cumul)', 8420, 9870, '-14.7%'],
    ['Disponibilite Reseau (%)', '99.2%', '98.8%', '+0.4 pts'],
    ['Interventions Terrain', 298, 274, '+8.8%'],
    ['Cout Total Interventions (DA)', '4 218 500', '3 985 000', '+5.9%'],
]
for i, row in enumerate(data_rows, 5):
    ws1.append(row)
    fill = alt_fill if i % 2 == 0 else PatternFill(fill_type=None)
    for col in range(1, 5):
        cell = ws1.cell(row=i, column=col)
        if fill.fill_type:
            cell.fill = fill
        cell.border = thin
        cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')

# Sheet 2: Detail par Wilaya
ws2 = wb.create_sheet('Detail par Wilaya')
ws2.column_dimensions['A'].width = 20
for col in ['B', 'C', 'D', 'E', 'F']:
    ws2.column_dimensions[col].width = 18

ws2.append(['Wilaya', 'Nb Incidents', 'Critiques', 'MTTR Moy. (h)', 'Clients Impact.', 'Taux Resol. (%)'])
for col in range(1, 7):
    cell = ws2.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin

wilaya_data = [
    ['Alger', 89, 7, 3.8, 2840, 92.1],
    ['Oran', 54, 3, 4.1, 1290, 90.7],
    ['Constantine', 42, 2, 4.6, 980, 88.1],
    ['Annaba', 38, 1, 4.9, 870, 89.5],
    ['Blida', 31, 2, 5.2, 720, 87.1],
    ['Setif', 28, 1, 4.3, 640, 91.0],
    ['Tlemcen', 19, 1, 5.8, 420, 84.2],
    ['Bejaia', 11, 1, 4.4, 660, 90.9],
]
for i, row in enumerate(wilaya_data, 2):
    ws2.append(row)
    fill = alt_fill if i % 2 == 0 else PatternFill(fill_type=None)
    for col in range(1, 7):
        cell = ws2.cell(row=i, column=col)
        if fill.fill_type:
            cell.fill = fill
        cell.border = thin
        cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')

xlsx_path = os.path.join(SANDBOX, 'rapport_incidents_mensuel.xlsx')
wb.save(xlsx_path)
print(f"  [OK] rapport_incidents_mensuel.xlsx (2 feuilles)")

# --- 2d: pipeline_etl_config.yaml ---
yaml_content = """# SOTIFibre BI Platform - Configuration Pipeline ETL
# Pipeline : ETL_Incidents_Reseau
# Version  : 3.2
# Genere   : 2026-05-15

pipeline:
  name: ETL_Incidents_Reseau
  description: >
    Extraction des incidents reseau depuis la base PostgreSQL de production,
    normalisation des severites, enrichissement geographique et chargement
    dans la table de faits fact_incidents_reseau du Data Warehouse.
  version: "3.2"
  category: Reseau
  priority: 1

schedule:
  enabled: true
  cron: "0 2 * * *"
  timezone: Europe/Paris
  retry_on_failure: true
  max_retries: 3
  retry_delay_seconds: 300

source:
  name: SRC_Production_PostgreSQL
  type: postgresql
  host: 192.168.224.128
  port: 5432
  database: sotifibre_db
  schema: public
  table: incidents_reseau
  username: sotifibre_readonly
  fetch_mode: incremental
  incremental_column: date_signalement
  batch_size: 5000
  connection_timeout: 30

transformations:
  - step: 1
    name: Filtrage doublons
    type: deduplication
    key_columns: [id, date_signalement]
    action: keep_latest

  - step: 2
    name: Normalisation severite
    type: mapping
    column: severite
    mappings:
      "P1": "critique"
      "P2": "majeur"
      "P3": "mineur"
      "CRITIQUE": "critique"
      "MAJEUR": "majeur"
      "MINEUR": "mineur"
    default: "mineur"

  - step: 3
    name: Enrichissement geographique
    type: lookup
    lookup_table: dim_localisation
    join_on: zone_id
    add_columns: [wilaya, daira, commune, region]

  - step: 4
    name: Calcul duree coupure
    type: computed_column
    output_column: duree_coupure_h
    formula: "EXTRACT(EPOCH FROM (date_resolution - date_signalement)) / 3600.0"
    condition: "statut = 'resolu' AND date_resolution IS NOT NULL"

  - step: 5
    name: Filtre qualite donnees
    type: filter
    conditions:
      - column: date_signalement
        operator: not_null
      - column: severite
        operator: in
        values: [critique, majeur, mineur]
    on_reject: log_and_skip

target:
  name: fact_incidents_reseau
  schema: sotifibre_dw
  type: postgresql
  host: 192.168.224.128
  port: 5432
  database: supervisionolt_db
  load_strategy: upsert
  conflict_key: [id]
  truncate_before_load: false

notifications:
  on_success: true
  on_failure: true
  recipients:
    - admin@sotifibre.dz
    - dev.bi@sotifibre.dz
  channels: [email, platform]

quality_checks:
  min_rows: 1000
  max_null_pct_critical_cols: 5
  expected_columns: [id, type_incident, severite, date_signalement, zone_id]
  row_count_deviation_pct: 20
"""

yaml_path = os.path.join(SANDBOX, 'pipeline_etl_config.yaml')
with open(yaml_path, 'w', encoding='utf-8') as f:
    f.write(yaml_content)
print(f"  [OK] pipeline_etl_config.yaml")

# --- 2e: extract_clients_contrats.sql ---
sql_content = """-- ============================================================
-- SOTIFibre BI Platform - Script d'extraction SQL
-- Objectif : Extraction clients et contrats pour alimentation
--            de la dimension dim_client (SCD Type 2)
-- Source   : sotifibre_db.public
-- Cible    : dim_client (Data Warehouse)
-- Auteur   : Equipe Data Engineering SOTIFibre
-- Version  : 2.1 - 2026-05-15
-- ============================================================

-- 1. Vue intermediaire : Clients avec contrat actif
CREATE OR REPLACE VIEW v_clients_actifs AS
SELECT
    c.id                                        AS client_id,
    c.nom,
    c.prenom,
    c.email,
    c.telephone,
    c.wilaya,
    c.statut                                    AS statut_client,
    ct.id                                       AS contrat_id,
    ct.offre_id,
    o.libelle                                   AS libelle_offre,
    o.debit_descendant_mbps                     AS debit_nominal_mbps,
    o.tarif_mensuel_da,
    ct.date_debut                               AS date_debut_contrat,
    ct.date_fin                                 AS date_fin_contrat,
    ct.statut                                   AS statut_contrat,
    EXTRACT(YEAR FROM AGE(NOW(), c.date_activation))::INTEGER AS anciennete_annees,
    CASE
        WHEN ct.tarif_mensuel_da >= 5000 THEN 'Premium'
        WHEN ct.tarif_mensuel_da >= 3000 THEN 'Standard'
        ELSE 'Essentiel'
    END                                         AS segment_commercial
FROM clients c
INNER JOIN contrats ct ON ct.client_id = c.id
    AND ct.statut = 'actif'
    AND (ct.date_fin IS NULL OR ct.date_fin > NOW())
LEFT JOIN offres o ON o.id = ct.offre_id
WHERE c.statut = 'actif';

-- 2. Detection des changements pour SCD Type 2
WITH clients_actuels AS (
    SELECT
        client_id,
        MD5(CONCAT_WS('|', nom, prenom, email, wilaya, offre_id, statut_contrat)) AS hash_attributs
    FROM v_clients_actifs
),
clients_dw AS (
    SELECT
        code_naturel::INTEGER AS client_id,
        hash_attributs
    FROM dim_client
    WHERE est_courant = TRUE
)
SELECT
    ca.client_id,
    CASE
        WHEN cd.client_id IS NULL     THEN 'NOUVEAU'
        WHEN ca.hash_attributs != cd.hash_attributs THEN 'MODIFIE'
        ELSE 'INCHANGE'
    END AS action_requise
FROM clients_actuels ca
LEFT JOIN clients_dw cd ON cd.client_id = ca.client_id;

-- 3. Extraction finale pour chargement ETL
SELECT
    va.client_id,
    va.nom,
    va.prenom,
    va.email,
    va.telephone,
    va.wilaya,
    va.statut_client,
    va.contrat_id,
    va.offre_id,
    va.libelle_offre,
    va.debit_nominal_mbps,
    va.tarif_mensuel_da,
    va.date_debut_contrat,
    va.anciennete_annees,
    va.segment_commercial,
    NOW()                               AS date_extraction,
    'ETL_Clients_CSV_Dimension_v2.1'    AS source_pipeline
FROM v_clients_actifs va
ORDER BY va.client_id;

-- 4. Statistiques de controle post-extraction
SELECT
    COUNT(*)                            AS total_clients_actifs,
    COUNT(DISTINCT wilaya)              AS nb_wilayas,
    COUNT(DISTINCT offre_id)            AS nb_offres,
    SUM(tarif_mensuel_da)               AS revenu_mensuel_da,
    ROUND(AVG(anciennete_annees), 1)    AS anciennete_moyenne,
    MIN(date_debut_contrat)             AS premier_contrat,
    MAX(date_debut_contrat)             AS dernier_contrat
FROM v_clients_actifs;
"""

sql_path = os.path.join(SANDBOX, 'extract_clients_contrats.sql')
with open(sql_path, 'w', encoding='utf-8') as f:
    f.write(sql_content)
print(f"  [OK] extract_clients_contrats.sql")

# --- 2f: avatar_utilisateur.png (minimal valid PNG 64x64 gradient) ---
def create_png(filename, width=64, height=64):
    """Create a minimal valid PNG file (blue-navy gradient)."""
    def make_chunk(chunk_type, data):
        c = chunk_type + data
        crc = zlib.crc32(c) & 0xffffffff
        return struct.pack('>I', len(data)) + c + struct.pack('>I', crc)

    # IHDR
    ihdr_data = struct.pack('>IIBBBBB', width, height, 8, 2, 0, 0, 0)
    ihdr = make_chunk(b'IHDR', ihdr_data)

    # IDAT: build raw scanlines
    raw = bytearray()
    for y in range(height):
        raw.append(0)  # filter type: None
        for x in range(width):
            r = int(30 + (x / width) * 40)
            g = int(58 + (y / height) * 30)
            b = int(95 + (x / width) * 60)
            raw += bytes([r, g, b])

    compressed = zlib.compress(bytes(raw), 9)
    idat = make_chunk(b'IDAT', compressed)

    # IEND
    iend = make_chunk(b'IEND', b'')

    png_signature = b'\x89PNG\r\n\x1a\n'
    with open(filename, 'wb') as f:
        f.write(png_signature + ihdr + idat + iend)

png_path = os.path.join(SANDBOX, 'avatar_utilisateur.png')
create_png(png_path, 64, 64)
print(f"  [OK] avatar_utilisateur.png (64x64 px, navy gradient)")

# --- 2g: rapport_mensuel_performances.pdf ---
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

pdf_path = os.path.join(SANDBOX, 'rapport_mensuel_performances.pdf')
doc = SimpleDocTemplate(pdf_path, pagesize=A4,
                        rightMargin=2*cm, leftMargin=2*cm,
                        topMargin=2*cm, bottomMargin=2*cm)

styles = getSampleStyleSheet()
navy = colors.HexColor('#1E3A5F')
amber = colors.HexColor('#D4921A')

title_style = ParagraphStyle('Title', parent=styles['Title'],
    fontSize=18, textColor=navy, spaceAfter=6, alignment=TA_CENTER)
subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'],
    fontSize=10, textColor=colors.gray, alignment=TA_CENTER, spaceAfter=20)
section_style = ParagraphStyle('Section', parent=styles['Heading2'],
    fontSize=12, textColor=navy, spaceBefore=16, spaceAfter=8)
body_style = ParagraphStyle('Body', parent=styles['Normal'],
    fontSize=9, spaceAfter=6, leading=14)
footer_style = ParagraphStyle('Footer', parent=styles['Normal'],
    fontSize=8, textColor=colors.gray, alignment=TA_CENTER)

story = []

story.append(Paragraph('SOTIFibre - Rapport Mensuel Performances Reseau', title_style))
story.append(Paragraph('Periode : Mai 2026 | Genere le 15/05/2026 | Confidentiel', subtitle_style))
story.append(HRFlowable(width='100%', thickness=2, color=navy))
story.append(Spacer(1, 0.4*cm))

story.append(Paragraph('1. Synthese Executif', section_style))
story.append(Paragraph(
    'Au cours du mois de mai 2026, le reseau fibre SOTIFibre a maintenu un taux de disponibilite '
    'de <b>99,2%</b>, depassant l\'objectif SLA de 99,0%. Le nombre total d\'incidents enregistres '
    's\'etablit a <b>312</b>, en legere hausse (+8,7%) par rapport au mois precedent, '
    'principalement imputable a des conditions climatiques defavorables dans les wilayas du nord. '
    'Le MTTR moyen s\'ameliore de <b>17,6%</b> a 4,2 heures grace aux optimisations apportees '
    'au processus d\'escalade technique.', body_style))

story.append(Paragraph('2. Indicateurs Cles de Performance', section_style))
kpi_data = [
    ['Indicateur', 'Mai 2026', 'Objectif', 'Statut'],
    ['Disponibilite Reseau', '99,2 %', '99,0 %', 'Atteint'],
    ['MTTR Moyen', '4,2 h', '< 6,0 h', 'Atteint'],
    ['Taux Resolution < 24h', '91,2 %', '> 90,0 %', 'Atteint'],
    ['Incidents Critiques Ouverts', '3', '< 5', 'Atteint'],
    ['Bande Passante Utilisee Moy.', '73,4 %', '< 85 %', 'Atteint'],
    ['Clients Actifs Fibre', '42 380', '45 000', 'En cours'],
    ['Satisfaction Client', '87,3 %', '> 90,0 %', 'A ameliorer'],
]
ts = TableStyle([
    ('BACKGROUND', (0, 0), (-1, 0), navy),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('FONTSIZE', (0, 0), (-1, -1), 8),
    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EEF2F7')]),
    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
    ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ('TOPPADDING', (0, 0), (-1, -1), 5),
    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
])
kpi_table = Table(kpi_data, colWidths=[6*cm, 3*cm, 3*cm, 3*cm])
kpi_table.setStyle(ts)
story.append(kpi_table)

story.append(Paragraph('3. Analyse des Incidents par Region', section_style))
inc_data = [
    ['Wilaya', 'Total', 'Critiques', 'MTTR (h)', 'Clients Impact.'],
    ['Alger', '89', '7', '3,8', '2 840'],
    ['Oran', '54', '3', '4,1', '1 290'],
    ['Constantine', '42', '2', '4,6', '980'],
    ['Annaba', '38', '1', '4,9', '870'],
    ['Blida', '31', '2', '5,2', '720'],
    ['Autres wilayas', '58', '3', '4,8', '1 720'],
    ['TOTAL', '312', '18', '4,2', '8 420'],
]
ts2 = TableStyle([
    ('BACKGROUND', (0, 0), (-1, 0), navy),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D4E0F0')),
    ('FONTSIZE', (0, 0), (-1, -1), 8),
    ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#EEF2F7')]),
    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
    ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
    ('TOPPADDING', (0, 0), (-1, -1), 5),
    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
])
inc_table = Table(inc_data, colWidths=[5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 3.5*cm])
inc_table.setStyle(ts2)
story.append(inc_table)

story.append(Paragraph('4. Recommandations', section_style))
recommandations = [
    '<b>Satisfaction client :</b> Lancer une enquete de satisfaction ciblee dans les wilayas '
    'avec MTTR superieur a 5h (Blida, Annaba) pour identifier les axes d\'amelioration prioritaires.',
    '<b>Capacite reseau :</b> Les noeuds OLT-ALGER-03 et OLT-BLIDA-01 approchent du seuil '
    'de 85% d\'utilisation. Planifier une extension de capacite avant la fin du T2 2026.',
    '<b>Recrutement clients :</b> L\'objectif 45 000 abonnes montre un ecart de 2 620 unites. '
    'Intensifier les actions commerciales dans les wilayas sous-representees.',
]
for rec in recommandations:
    story.append(Paragraph(f'• {rec}', body_style))

story.append(Spacer(1, 0.5*cm))
story.append(HRFlowable(width='100%', thickness=0.5, color=colors.gray))
story.append(Spacer(1, 0.2*cm))
story.append(Paragraph(
    'Document confidentiel - SOTIFibre Direction des Systemes d\'Information | '
    'Genere automatiquement par la Plateforme BI Integrated BI | Mai 2026',
    footer_style))

doc.build(story)
print(f"  [OK] rapport_mensuel_performances.pdf (4 sections)")

# ============================================================
# STEP 3: Print summary
# ============================================================
print("\n" + "=" * 60)
print("  SANDBOX CREE - test_files_sandbox/")
print("=" * 60)
files = sorted(os.listdir(SANDBOX))
for fname in files:
    fpath = os.path.join(SANDBOX, fname)
    size = os.path.getsize(fpath)
    unit = 'KB' if size > 1024 else 'B'
    val = size // 1024 if size > 1024 else size
    print(f"  {fname:<45} {val:>6} {unit}")
print("=" * 60)
print(f"\n[OK] seed_data.py mis a jour et sandbox genere dans :\n  {SANDBOX}")
